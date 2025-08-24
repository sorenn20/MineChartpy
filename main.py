# ============================================
#   HECHO POR MATÍAS ALFARO RUZ
# ============================================
import sys
import pandas as pd
import numpy as np
import pyvista as pv
import warnings
import openpyxl
from PyQt5 import QtWidgets, QtCore, QtGui
from pyvistaqt import QtInteractor
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from mplcursors import cursor
from vtkmodules.vtkCommonCore import vtkCommand
from vtkmodules.vtkRenderingCore import vtkCellPicker
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from scipy import stats
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
import json
import os
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import ChartLines


class CustomizationDialog(QtWidgets.QDialog):
    def __init__(self, current_settings, parent=None):
        super().__init__(parent)
        self.settings = current_settings.copy()
        self.setWindowTitle("Personalizar Gráfico")
        self.setup_ui()
        self.cargar_configuracion()

    def setup_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        self.line_style = QtWidgets.QComboBox()
        self.line_style.addItems(['-', '--', '-.', ':'])
        self.line_width = QtWidgets.QSpinBox()
        self.line_width.setRange(1, 10)
        self.font_size = QtWidgets.QSpinBox()
        self.font_size.setRange(8, 20)
        form = QtWidgets.QFormLayout()
        form.addRow("Estilo de línea:", self.line_style)
        form.addRow("Grosor de línea:", self.line_width)
        form.addRow("Tamaño de fuente:", self.font_size)
        layout.addLayout(form)
        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def cargar_configuracion(self):
        self.line_style.setCurrentText(self.settings.get('line_style', '-'))
        self.line_width.setValue(self.settings.get('line_width', 1))
        self.font_size.setValue(self.settings.get('font_size', 10))

    def get_settings(self):
        return {
            'line_style': self.line_style.currentText(),
            'line_width': self.line_width.value(),
            'tonnage_color': self.settings.get('tonnage_color', 'b'),
            'ley_color': self.settings.get('ley_color', 'r'),
            'xaxis_color': self.settings.get('xaxis_color', 'black'),
            'yaxis_color': self.settings.get('yaxis_color', 'black'),
            'font_size': self.font_size.value()
        }

class ColumnMapperDialog(QtWidgets.QDialog):
    def __init__(self, columns, parent=None):
        super().__init__(parent)
        self.columns = columns
        self.mappings = {}
        self.setWindowTitle("Mapear Columnas")
        self.setup_ui()

    def setup_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        self.required_columns = ['x0', 'y0', 'z0', 'xinc', 'yinc', 'zinc', 'ley']
        self.combos = {}
        form = QtWidgets.QFormLayout()
        for col in self.required_columns:
            combo = QtWidgets.QComboBox()
            combo.addItem("No asignar")
            combo.addItems(self.columns)
            self.combos[col] = combo
            form.addRow(f"{col}:", combo)
        layout.addLayout(form)
        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def accept(self):
        self.mappings = {col: self.combos[col].currentText() for col in self.required_columns if self.combos[col].currentText() != "No asignar"}
        if all(col in self.mappings for col in self.required_columns):
            super().accept()
        else:
            QtWidgets.QMessageBox.warning(self, "Error", "Debe asignar todas las columnas requeridas.")

class FormulaEditor(QtWidgets.QDialog):
    def __init__(self, df, file_path, parent=None):
        super().__init__(parent)
        self.df = df.copy()
        self.file_path = file_path
        self.setWindowTitle("Editor de Fórmulas")
        self.setup_ui()

    def setup_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        self.column_combo = QtWidgets.QComboBox()
        self.column_combo.addItems(self.df.columns)
        self.formula_input = QtWidgets.QLineEdit()
        self.formula_input.setPlaceholderText("Ejemplo: col1 * 2 + col2")
        btn_apply = QtWidgets.QPushButton("Aplicar Fórmula")
        btn_apply.clicked.connect(self.apply_formula)
        layout.addWidget(QtWidgets.QLabel("Seleccionar columna:"))
        layout.addWidget(self.column_combo)
        layout.addWidget(QtWidgets.QLabel("Fórmula:"))
        layout.addWidget(self.formula_input)
        layout.addWidget(btn_apply)
        self.table = QtWidgets.QTableWidget()
        self.update_table()
        layout.addWidget(self.table)
        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def update_table(self):
        self.table.setRowCount(len(self.df))
        self.table.setColumnCount(len(self.df.columns))
        self.table.setHorizontalHeaderLabels(self.df.columns)
        for i in range(len(self.df)):
            for j in range(len(self.df.columns)):
                item = QtWidgets.QTableWidgetItem(str(self.df.iloc[i, j]))
                self.table.setItem(i, j, item)

    def apply_formula(self):
        try:
            col = self.column_combo.currentText()
            formula = self.formula_input.text()
            if not formula:
                raise ValueError("Ingrese una fórmula")
            safe_dict = {c: self.df[c] for c in self.df.columns}
            safe_dict.update({'np': np})
            self.df[col] = eval(formula, {"__builtins__": {}}, safe_dict)
            self.update_table()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Error en la fórmula: {str(e)}")

class StatisticalReportDialog(QtWidgets.QDialog):
    def __init__(self, df, variable, parent=None):
        super().__init__(parent)
        self.df = df
        self.variable = variable
        self.setWindowTitle("Reporte Estadístico")
        self.setup_ui()

    def setup_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        stats_text = self.generate_stats()
        self.stats_label = QtWidgets.QLabel(stats_text)
        self.stats_label.setWordWrap(True)
        layout.addWidget(self.stats_label)
        self.figure = Figure()
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas)
        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok)
        buttons.accepted.connect(self.accept)
        layout.addWidget(buttons)
        self.plot_histogram()

    def generate_stats(self):
        data = self.df[self.variable].dropna()
        return (
            f"Variable: {self.variable}\n"
            f"Media: {data.mean():.2f}\n"
            f"Mediana: {data.median():.2f}\n"
            f"Desviación Estándar: {data.std():.2f}\n"
            f"Mínimo: {data.min():.2f}\n"
            f"Máximo: {data.max():.2f}\n"
            f"Conteo: {len(data)}"
        )

    def plot_histogram(self):
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        data = self.df[self.variable].dropna()
        ax.hist(data, bins=30, color='blue', alpha=0.7)
        ax.set_title(f"Histograma de {self.variable}")
        ax.set_xlabel(self.variable)
        ax.set_ylabel("Frecuencia")
        self.canvas.draw()

class TonnageLeyDialog(QtWidgets.QDialog):
    def __init__(self, df, parent=None):
        super().__init__(parent)
        self.mpl_cursor = None
        self.last_calculation = ()
        self.is_exporting = False
        self.setWindowTitle("Curva Tonelaje-Ley")
        self.setGeometry(100, 100, 1000, 800)
        self.df = df
        self.custom_settings = {
            'line_style': '-', 'line_width': 1, 'tonnage_color': 'b',
            'ley_color': 'r', 'xaxis_color': 'black', 'yaxis_color': 'black',
            'font_size': 10
        }
        self.scenarios = []
        self.setup_ui()

    def setup_ui(self):
        main_layout = QtWidgets.QVBoxLayout(self)
        grid_layout = QtWidgets.QGridLayout()
        self.combos = {
            'x': QtWidgets.QComboBox(), 'y': QtWidgets.QComboBox(),
            'z': QtWidgets.QComboBox(), 'ley': QtWidgets.QComboBox()
        }
        labels = [
            ("Tamaño X:", 'x', 0), ("Tamaño Y:", 'y', 1),
            ("Tamaño Z:", 'z', 2), ("Ley (%):", 'ley', 3)
        ]
        numeric_cols = self.df.select_dtypes(include=np.number).columns.tolist()
        for combo in self.combos.values():
            combo.addItems(numeric_cols)
        for text, key, row in labels:
            grid_layout.addWidget(QtWidgets.QLabel(text), row, 0)
            grid_layout.addWidget(self.combos[key], row, 1)
        params_layout = QtWidgets.QHBoxLayout()
        self.densidad = QtWidgets.QLineEdit()
        self.incremento = QtWidgets.QLineEdit("0.01")
        params_layout.addWidget(QtWidgets.QLabel("Densidad (t/m³):"))
        params_layout.addWidget(self.densidad)
        params_layout.addWidget(QtWidgets.QLabel("Incremento (%):"))
        params_layout.addWidget(self.incremento)
        self.btn_calcular = QtWidgets.QPushButton("Calcular")
        self.btn_calcular.clicked.connect(self.calcular)
        reserve_layout = QtWidgets.QHBoxLayout()
        self.cutoff_input = QtWidgets.QLineEdit()
        self.btn_calcular_reserva = QtWidgets.QPushButton("Calcular Reserva")
        self.btn_save_scenario = QtWidgets.QPushButton("Guardar Escenario")
        self.btn_save_scenario.clicked.connect(self.save_scenario)
        self.btn_compare_scenarios = QtWidgets.QPushButton("Comparar Escenarios")
        self.btn_compare_scenarios.clicked.connect(self.compare_scenarios)
        self.reserve_result = QtWidgets.QLabel()
        self.reserve_result.setWordWrap(True)
        reserve_layout.addWidget(QtWidgets.QLabel("Ley de Corte (%):"))
        reserve_layout.addWidget(self.cutoff_input)
        reserve_layout.addWidget(self.btn_calcular_reserva)
        reserve_layout.addWidget(self.btn_save_scenario)
        reserve_layout.addWidget(self.btn_compare_scenarios)
        self.figure = Figure()
        self.canvas = FigureCanvas(self.figure)
        self.ax = self.figure.add_subplot(111)
        self.ax2 = self.ax.twinx()
        self.toolbar = NavigationToolbar(self.canvas, self)
        self.btn_exportar = QtWidgets.QPushButton("Exportar Gráfico")
        self.btn_exportar.clicked.connect(self.exportar_grafico)
        self.btn_export_excel = QtWidgets.QPushButton("Exportar a Excel")
        self.btn_export_excel.clicked.connect(self.export_to_excel)
        self.btn_limpiar = QtWidgets.QPushButton("Limpiar Referencias")
        self.btn_limpiar.clicked.connect(self.limpiar_referencias)
        self.btn_personalizar = QtWidgets.QPushButton("Personalizar Gráfico")
        self.btn_personalizar.clicked.connect(self.abrir_personalizacion)
        main_layout.addWidget(self.btn_personalizar)
        axis_color_layout = QtWidgets.QHBoxLayout()
        self.btn_xaxis_color = QtWidgets.QPushButton("Color Título Eje X")
        self.btn_yaxis_color = QtWidgets.QPushButton("Color Título Eje Y")
        self.btn_xaxis_color.clicked.connect(self.cambiar_color_xaxis)
        self.btn_yaxis_color.clicked.connect(self.cambiar_color_yaxis)
        axis_color_layout.addWidget(self.btn_xaxis_color)
        axis_color_layout.addWidget(self.btn_yaxis_color)
        main_layout.addLayout(grid_layout)
        main_layout.addLayout(params_layout)
        main_layout.addWidget(self.btn_calcular)
        main_layout.addLayout(reserve_layout)
        main_layout.addLayout(axis_color_layout)
        main_layout.addWidget(self.reserve_result)
        main_layout.addWidget(self.toolbar)
        main_layout.addWidget(self.canvas)
        main_layout.addWidget(self.btn_exportar)
        main_layout.addWidget(self.btn_export_excel)
        main_layout.addWidget(self.btn_limpiar)
        self.btn_calcular_reserva.clicked.connect(self.calcular_reserva)
        self.canvas.mpl_connect('button_press_event', self.on_click)

    def on_click(self, event):
        if event.inaxes is not None:
            self.ax.axvline(x=event.xdata, color='gray', linestyle='--', alpha=0.7)
            self.canvas.draw()

    def exportar_grafico(self):
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Exportar Gráfico", "", 
            "PNG (*.png);;JPEG (*.jpg);;PDF (*.pdf)"
        )
        if file_path:
            try:
                if hasattr(self, 'mpl_cursor'):
                    self.mpl_cursor.remove()
                self.figure.savefig(file_path, bbox_inches='tight')
                if hasattr(self, 'mpl_cursor'):
                    self.mpl_cursor = cursor(self.ax, hover=True)
                    self.mpl_cursor.connect("add", lambda sel: sel.annotation.set_text(
                        f"Ley: {cortes[int(sel.index)]:.2f}%\n"
                        f"Tonelaje: {tonelajes[int(sel.index)]:.2f} Mt\n"
                        f"Ley Media: {leyes_medias[int(sel.index)]:.2f}%"
                    ))
            except Exception as e:
                QtWidgets.QMessageBox.critical(
                    self, "Error al exportar", f"No se pudo guardar el gráfico:\n{str(e)}"
                )

    def export_to_excel(self):
        if not self.last_calculation:
            QtWidgets.QMessageBox.warning(self, "Advertencia", "Realice un cálculo primero.")
            return

        try:
            cortes, tonelajes, leyes_medias = self.last_calculation
            print(f"[DEPURACIÓN] Longitud de cortes: {len(cortes)}, tonelajes: {len(tonelajes)}, leyes_medias: {len(leyes_medias)}")
            print(f"[DEPURACIÓN] Primeros 5 valores - cortes: {cortes[:5]}, tonelajes: {tonelajes[:5]}, leyes_medias: {leyes_medias[:5]}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"No se pudo acceder a los datos del cálculo: {str(e)}")
            return

        if len(cortes) == 0 or len(cortes) != len(tonelajes) or len(cortes) != len(leyes_medias):
            QtWidgets.QMessageBox.critical(self, "Error", "Datos inválidos o inconsistentes.")
            return

        try:
            densidad = float(self.densidad.text())
            incremento = float(self.incremento.text())
            x_col = self.combos['x'].currentText()
            y_col = self.combos['y'].currentText()
            z_col = self.combos['z'].currentText()
            ley_col = self.combos['ley'].currentText()
        except ValueError:
            QtWidgets.QMessageBox.critical(self, "Error", "Densidad e incremento deben ser valores numéricos.")
            return

        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Exportar a Excel", "", "Excel (*.xlsx)")
        if not file_path:
            return

        self.setEnabled(False)
        QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Curva Tonelaje-Ley"

            # Escribir parámetros de entrada
            ws['A1'] = "Parámetros de Entrada"
            ws['A1'].font = openpyxl.styles.Font(bold=True)
            params = [
                ("Densidad (t/m³)", densidad),
                ("Incremento (%)", incremento),
                ("Columna X", x_col),
                ("Columna Y", y_col),
                ("Columna Z", z_col),
                ("Columna Ley", ley_col)
            ]
            for i, (label, value) in enumerate(params, start=2):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value

            # Escribir datos de la curva
            start_row = 10
            headers = ["Ley de Corte (%)", "Tonelaje (Mt)", "Ley Media (%)"]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=start_row, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)

            for i, (corte, tonelaje, ley_media) in enumerate(zip(cortes, tonelajes, leyes_medias), start=start_row + 1):
                ws.cell(row=i, column=1, value=float(corte))
                ws.cell(row=i, column=2, value=float(tonelaje))
                ws.cell(row=i, column=3, value=float(ley_media))

            print(f"[DEPURACIÓN] Datos escritos desde fila {start_row + 1} hasta {start_row + len(cortes)}")

            # Crear gráfico de líneas
            chart = LineChart()
            chart.title = "Curva Tonelaje-Ley"
            chart.width = 15
            chart.height = 10

            # Definir categorías (Ley de Corte) como eje X
            cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(cortes))
            chart.set_categories(cats)
            chart.x_axis.title = "Ley de Corte (%)"
            chart.x_axis.number_format = '0.00%'  
            print(f"[DEPURACIÓN] Categorías definidas: min_col=1, min_row={start_row + 1}, max_row={start_row + len(cortes)}")

            # Serie 1: Tonelaje (eje Y primario)
            data_ton = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(cortes))
            ton_series = Series(data_ton, title_from_data=True)
            chart.append(ton_series)  
            chart.y_axis.title = "Tonelaje (Millones t)"
            chart.y_axis.majorGridlines = None  
            print(f"[DEPURACIÓN] Serie Tonelaje añadida: min_col=2, min_row={start_row}, max_row={start_row + len(cortes)}")

            # Serie 2: Ley Media (eje Y secundario)
            data_ley = Reference(ws, min_col=3, min_row=start_row, max_row=start_row + len(cortes))
            ley_series = Series(data_ley, title_from_data=True)
            ley_series.axis = "secondary"  
            chart.append(ley_series)
            # Configurar el eje secundario
            chart._charts.append(LineChart())  
            chart.y_axis2 = chart._charts[1].y_axis  
            chart.y_axis2.title = "Ley Media (%)"
            chart.y_axis2.number_format = '0.00%'
            chart.y_axis2.crosses = "max"
            print(f"[DEPURACIÓN] Serie Ley Media añadida: min_col=3, min_row={start_row}, max_row={start_row + len(cortes)}")

            # Añadir gráfico al worksheet
            ws.add_chart(chart, "E2")
            print("[DEPURACIÓN] Gráfico añadido en E2")

            # Ajustar ancho de columnas
            for col in ['A', 'B', 'C']:
                ws.column_dimensions[col].width = 15

            # Guardar el archivo
            wb.save(file_path)
            print(f"[DEPURACIÓN] Archivo guardado en: {file_path}")
            QtWidgets.QMessageBox.information(self, "Éxito", "Curva exportada a Excel con gráfico!")

        except Exception as e:
            print(f"[DEPURACIÓN] Error durante la exportación: {str(e)}")
            QtWidgets.QMessageBox.critical(self, "Error", f"No se pudo exportar:\n{str(e)}")
        finally:
            self.setEnabled(True)
            QtWidgets.QApplication.restoreOverrideCursor()

    def limpiar_referencias(self):
        for line in self.ax.lines:
            if line.get_linestyle() == '--':
                line.remove()
        self.canvas.draw()

    def calcular_reserva(self):
        try:
            if not all(combo.currentText() for combo in self.combos.values()):
                raise ValueError("Debe seleccionar todas las columnas requeridas")
            if not self.densidad.text() or not self.cutoff_input.text():
                raise ValueError("Complete todos los parámetros de cálculo")
            cutoff = float(self.cutoff_input.text())
            densidad = float(self.densidad.text())
            x_col = self.combos['x'].currentText()
            y_col = self.combos['y'].currentText()
            z_col = self.combos['z'].currentText()
            ley_col = self.combos['ley'].currentText()
            self.df['tonelaje_temp'] = (
                self.df[x_col].astype(float) * self.df[y_col].astype(float) *
                self.df[z_col].astype(float) * densidad
            )
            subset = self.df[self.df[ley_col].astype(float) >= cutoff]
            ton_acum = subset['tonelaje_temp'].sum()
            ley_media = (subset[ley_col].astype(float) * subset['tonelaje_temp']).sum() / ton_acum if ton_acum > 0 else 0
            self.reserve_result.setText(
                f"<b>Resultados para Ley de Corte {cutoff:.2f}%:</b>\n"
                f"Tonelaje: {ton_acum / 1e6:.2f} Mt\n"
                f"Ley Media: {ley_media:.2f}%"
            )
        except ValueError as ve:
            QtWidgets.QMessageBox.warning(self, "Error de validación", str(ve))
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Error en cálculo de reserva: {str(e)}")

    def save_scenario(self):
        try:
            if not all(combo.currentText() for combo in self.combos.values()):
                raise ValueError("Debe seleccionar todas las columnas requeridas")
            if not self.densidad.text() or not self.cutoff_input.text():
                raise ValueError("Complete todos los parámetros de cálculo")
            cutoff = float(self.cutoff_input.text())
            densidad = float(self.densidad.text())
            x_col = self.combos['x'].currentText()
            y_col = self.combos['y'].currentText()
            z_col = self.combos['z'].currentText()
            ley_col = self.combos['ley'].currentText()
            self.df['tonelaje_temp'] = (
                self.df[x_col].astype(float) * self.df[y_col].astype(float) *
                self.df[z_col].astype(float) * densidad
            )
            subset = self.df[self.df[ley_col].astype(float) >= cutoff]
            ton_acum = subset['tonelaje_temp'].sum()
            ley_media = (subset[ley_col].astype(float) * subset['tonelaje_temp']).sum() / ton_acum if ton_acum > 0 else 0
            scenario = {
                "cutoff": cutoff,
                "densidad": densidad,
                "tonelaje": ton_acum / 1e6,
                "ley_media": ley_media
            }
            self.scenarios.append(scenario)
            self.parent().tonelaje_scenarios.append(scenario)
            QtWidgets.QMessageBox.information(self, "Éxito", "Escenario guardado!")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Error al guardar escenario:\n{str(e)}")

    def compare_scenarios(self):
        if not self.scenarios:
            QtWidgets.QMessageBox.warning(self, "Advertencia", "No hay escenarios guardados para comparar.")
            return
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Comparar Escenarios")
        layout = QtWidgets.QVBoxLayout(dialog)
        table = QtWidgets.QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Ley de Corte (%)", "Densidad (t/m³)", "Tonelaje (Mt)", "Ley Media (%)"])
        table.setRowCount(len(self.scenarios))
        for i, scenario in enumerate(self.scenarios):
            table.setItem(i, 0, QtWidgets.QTableWidgetItem(f"{scenario['cutoff']:.2f}"))
            table.setItem(i, 1, QtWidgets.QTableWidgetItem(f"{scenario['densidad']:.2f}"))
            table.setItem(i, 2, QtWidgets.QTableWidgetItem(f"{scenario['tonelaje']:.2f}"))
            table.setItem(i, 3, QtWidgets.QTableWidgetItem(f"{scenario['ley_media']:.2f}"))
        layout.addWidget(table)
        dialog.exec_()

    def calcular(self):
        if self.is_exporting:
            print("[DEPURACIÓN] Cálculo bloqueado: exportación en progreso")
            QtWidgets.QMessageBox.warning(self, "Advertencia", "Espere a que finalice la exportación antes de realizar un nuevo cálculo.")
            return

        try:
            print("[DEPURACIÓN] Iniciando cálculo")
            self.btn_calcular.setEnabled(False)
            QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)

            if not all(combo.currentText() for combo in self.combos.values()):
                raise ValueError("Debe seleccionar todas las columnas requeridas")
            if not self.densidad.text() or not self.incremento.text():
                raise ValueError("Complete todos los parámetros de cálculo")
            densidad = float(self.densidad.text())
            if densidad <= 0:
                raise ValueError("La densidad debe ser mayor que 0")
            incremento = float(self.incremento.text())
            if incremento <= 0:
                raise ValueError("El incremento debe ser mayor que 0")
            x_col = self.combos['x'].currentText()
            y_col = self.combos['y'].currentText()
            z_col = self.combos['z'].currentText()
            ley_col = self.combos['ley'].currentText()

            for col in [x_col, y_col, z_col, ley_col]:
                if not pd.api.types.is_numeric_dtype(self.df[col]):
                    raise ValueError(f"La columna {col} debe contener datos numéricos")

            self.df['tonelaje'] = (
                self.df[x_col].astype(float) * self.df[y_col].astype(float) *
                self.df[z_col].astype(float) * densidad
            )
            cortes = []
            tonelajes = []
            leyes_medias = []
            current = self.df[ley_col].astype(float).min()
            max_ley = self.df[ley_col].astype(float).max()
            if pd.isna(current) or pd.isna(max_ley):
                raise ValueError("No se encontraron valores válidos en la columna de ley")
            while current <= max_ley:
                subset = self.df[self.df[ley_col].astype(float) >= current]
                ton_acum = subset['tonelaje'].sum()
                ley_media = (subset[ley_col].astype(float) * subset['tonelaje']).sum() / ton_acum if ton_acum > 0 else 0
                cortes.append(current)
                tonelajes.append(ton_acum / 1e6)
                leyes_medias.append(ley_media)
                current += incremento
            if not cortes:
                raise ValueError("No se generaron datos para la curva. Verifique los datos de entrada.")

            self.last_calculation = (cortes, tonelajes, leyes_medias)
            print(f"[DEPURACIÓN] self.last_calculation asignado: {self.last_calculation}")

            self.figure.clear()
            self.ax = self.figure.add_subplot(111)
            self.ax2 = self.ax.twinx()
            line1, = self.ax.plot(
                cortes, tonelajes, linestyle=self.custom_settings['line_style'],
                linewidth=self.custom_settings['line_width'], color=self.custom_settings['tonnage_color'],
                label='Tonelaje'
            )
            line2, = self.ax2.plot(
                cortes, leyes_medias, linestyle=self.custom_settings['line_style'],
                linewidth=self.custom_settings['line_width'], color=self.custom_settings['ley_color'],
                label='Ley Media'
            )
            self.ax.set_xlabel("Ley de Corte (%)", fontsize=self.custom_settings['font_size'])
            self.ax.set_ylabel("Tonelaje (Millones t)", color=self.custom_settings['tonnage_color'],
                            fontsize=self.custom_settings['font_size'])
            self.ax2.set_ylabel("Ley Media (%)", color=self.custom_settings['ley_color'],
                                fontsize=self.custom_settings['font_size'], rotation=270, va='bottom', labelpad=20)
            self.actualizar_colores_ejes()
            if tonelajes:
                self.ax.set_ylim(0, max(tonelajes) * 1.05)
            if leyes_medias:
                self.ax2.set_ylim(0, max(leyes_medias) * 1.05)
            lines = [line1, line2]
            labels = [l.get_label() for l in lines]
            self.ax.legend(lines, labels, loc='upper right')
            if hasattr(self, 'mpl_cursor') and self.mpl_cursor is not None:
                self.mpl_cursor.remove()
            self.mpl_cursor = cursor(self.ax, hover=True)
            self.mpl_cursor.connect("add", lambda sel: sel.annotation.set_text(
                f"Ley: {cortes[int(sel.index)]:.2f}%\n"
                f"Tonelaje: {tonelajes[int(sel.index)]:.2f} Mt\n"
                f"Ley Media: {leyes_medias[int(sel.index)]:.2f}%"
            ))
            self.canvas.draw()
            print(f"[DEPURACIÓN] Gráfico actualizado correctamente")

        except ValueError as ve:
            print(f"[DEPURACIÓN] Error en calcular (ValueError): {str(ve)}")
            QtWidgets.QMessageBox.warning(self, "Error de validación", str(ve))
            self.last_calculation = ()  
        except Exception as e:
            print(f"[DEPURACIÓN] Error inesperado en calcular: {str(e)}")
            QtWidgets.QMessageBox.critical(self, "Error inesperado", f"Ocurrió un error: {str(e)}")
            self.last_calculation = ()  
        finally:
            self.btn_calcular.setEnabled(True)
            QtWidgets.QApplication.restoreOverrideCursor()
            print("[DEPURACIÓN] Finalizando cálculo")
         

    def actualizar_colores_ejes(self):
        if hasattr(self, 'ax'):
            self.ax.xaxis.label.set_color(self.custom_settings['xaxis_color'])
            self.ax.yaxis.label.set_color(self.custom_settings['yaxis_color'])
            self.ax.tick_params(axis='x', colors=self.custom_settings['xaxis_color'])
            self.ax.tick_params(axis='y', colors=self.custom_settings['yaxis_color'])
            self.ax2.yaxis.label.set_color(self.custom_settings['yaxis_color'])
            self.ax2.tick_params(axis='y', colors=self.custom_settings['yaxis_color'])
            self.canvas.draw()

    def cambiar_color_xaxis(self):
        color = QtWidgets.QColorDialog.getColor()
        if color.isValid():
            self.custom_settings['xaxis_color'] = color.name()
            self.actualizar_colores_ejes()

    def cambiar_color_yaxis(self):
        color = QtWidgets.QColorDialog.getColor()
        if color.isValid():
            self.custom_settings['yaxis_color'] = color.name()
            self.actualizar_colores_ejes()

    def abrir_personalizacion(self):
        dialog = CustomizationDialog(self.custom_settings, self)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            self.custom_settings.update(dialog.get_settings())
            self.calcular()

class EconomicAnalysisDialog(QtWidgets.QDialog):
    def __init__(self, df, parent=None):
        super().__init__(parent)
        self.df = df
        self.average_grade = None
        self.setWindowTitle("Análisis Económico")
        self.setGeometry(100, 100, 1200, 800)
        self.setup_ui()
        self.load_default_values()

    def setup_ui(self):
        main_layout = QtWidgets.QVBoxLayout(self)
        self.tabs = QtWidgets.QTabWidget()
        self.setup_cost_tab()
        self.setup_processing_tab()
        self.setup_price_tab()
        self.setup_analysis_tab()
        self.setup_risk_tab()
        main_layout.addWidget(self.tabs)
        self.btn_export_excel = QtWidgets.QPushButton("Exportar a Excel")
        self.btn_export_excel.clicked.connect(self.export_to_excel)
        main_layout.addWidget(self.btn_export_excel)

    def setup_cost_tab(self):
        tab = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout()
        self.cost_table = QtWidgets.QTableWidget()
        self.cost_table.setColumnCount(7)
        self.cost_table.setHorizontalHeaderLabels([
            "Método", "Preparación", "Arranque", "Extracción", 
            "Relleno", "Transporte", "Total"
        ])
        layout.addWidget(QtWidgets.QLabel("Costos Mineros (US$/ton):"))
        layout.addWidget(self.cost_table)
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Costos Mina")

    def setup_processing_tab(self):
        tab = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout()
        self.mineral_type = QtWidgets.QComboBox()
        self.mineral_type.addItems([
            "Metales pesados", 
            "Minerales auríferos", 
            "Minerales industriales"
        ])
        self.planta_cost = QtWidgets.QLineEdit()
        self.recuperacion_mineral = QtWidgets.QLineEdit()
        self.recuperacion_metalurgica = QtWidgets.QLineEdit()
        self.maquila = QtWidgets.QLineEdit()
        form = QtWidgets.QFormLayout()
        form.addRow("Tipo de mineral:", self.mineral_type)
        form.addRow("Costo planta (US$/t):", self.planta_cost)
        form.addRow("Recuperación mineral (%):", self.recuperacion_mineral)
        form.addRow("Recuperación metalúrgica (%):", self.recuperacion_metalurgica)
        form.addRow("Maquila (%):", self.maquila)
        layout.addLayout(form)
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Procesamiento")

    def setup_price_tab(self):
        tab = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout()
        self.price_table = QtWidgets.QTableWidget()
        self.price_table.setColumnCount(2)
        self.price_table.setHorizontalHeaderLabels(["Metal", "Precio (USD)"])
        layout.addWidget(QtWidgets.QLabel("Precios de Metales:"))
        layout.addWidget(self.price_table)
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Precios")

    def setup_analysis_tab(self):
        tab = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout()
        self.selected_method = QtWidgets.QComboBox()
        self.costo_otros = QtWidgets.QLineEdit()
        self.production = QtWidgets.QLineEdit()
        self.discount_rate = QtWidgets.QLineEdit("10")
        self.project_years = QtWidgets.QLineEdit("10")
        self.inversion_inicial = QtWidgets.QLineEdit("1000000")
        self.G_and_A_input = QtWidgets.QLineEdit("0")
        self.depreciation_input = QtWidgets.QLineEdit("0")
        self.sustaining_capital_input = QtWidgets.QLineEdit("0")
        self.exploration_expenses_input = QtWidgets.QLineEdit("0")
        self.scenario_combo = QtWidgets.QComboBox()
        self.scenario_combo.currentTextChanged.connect(self.set_scenario)
        self.cutoff_grade_result = QtWidgets.QLabel("N/A")
        self.van_result = QtWidgets.QLabel("N/A")
        self.AISC_result = QtWidgets.QLabel("N/A")
        btn_calculate = QtWidgets.QPushButton("Calcular")
        btn_calculate.clicked.connect(self.calculate_values)

        form = QtWidgets.QFormLayout()
        form.addRow("Método de explotación:", self.selected_method)
        form.addRow("Costo otros (%):", self.costo_otros)
        form.addRow("Producción anual (ton):", self.production)
        form.addRow("Tasa de descuento (%):", self.discount_rate)
        form.addRow("Años del proyecto:", self.project_years)
        form.addRow("Inversión inicial (US$):", self.inversion_inicial)
        form.addRow("G&A Costs (US$ anuales):", self.G_and_A_input)
        form.addRow("Depreciación (US$ anuales):", self.depreciation_input)
        form.addRow("Capital sostenible (US$ anuales):", self.sustaining_capital_input)
        form.addRow("Gastos exploración (US$ anuales):", self.exploration_expenses_input)
        form.addRow("Seleccionar escenario:", self.scenario_combo)
        form.addRow("Ley de corte calculada:", self.cutoff_grade_result)
        form.addRow("VAN del proyecto:", self.van_result)
        form.addRow("AISC (US$/ton metal):", self.AISC_result)

        layout.addLayout(form)
        layout.addWidget(btn_calculate)
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Análisis")

    def setup_risk_tab(self):
        tab = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout()
        
        self.risk_params = {
            "Precio Metal": {"min": QtWidgets.QLineEdit("1800"), "mode": QtWidgets.QLineEdit("2000"), "max": QtWidgets.QLineEdit("2200")},
            "Costo Minero": {"min": QtWidgets.QLineEdit("8"), "mode": QtWidgets.QLineEdit("10"), "max": QtWidgets.QLineEdit("12")},
            "Costo Planta": {"min": QtWidgets.QLineEdit("4"), "mode": QtWidgets.QLineEdit("5"), "max": QtWidgets.QLineEdit("6")},
            "Producción Anual": {"min": QtWidgets.QLineEdit("900000"), "mode": QtWidgets.QLineEdit("1000000"), "max": QtWidgets.QLineEdit("1100000")},
            "Recuperación Total": {"min": QtWidgets.QLineEdit("85"), "mode": QtWidgets.QLineEdit("90"), "max": QtWidgets.QLineEdit("95")}
        }
        
        risk_form = QtWidgets.QFormLayout()
        for param, fields in self.risk_params.items():
            h_layout = QtWidgets.QHBoxLayout()
            h_layout.addWidget(QtWidgets.QLabel("Mín:"))
            h_layout.addWidget(fields["min"])
            h_layout.addWidget(QtWidgets.QLabel("Más Probable:"))
            h_layout.addWidget(fields["mode"])
            h_layout.addWidget(QtWidgets.QLabel("Máx:"))
            h_layout.addWidget(fields["max"])
            risk_form.addRow(param, h_layout)
        
        self.num_simulations = QtWidgets.QLineEdit("1000")
        risk_form.addRow("Número de Simulaciones:", self.num_simulations)
        
        btn_run_montecarlo = QtWidgets.QPushButton("Ejecutar Monte Carlo")
        btn_run_montecarlo.clicked.connect(self.run_montecarlo)
        
        self.montecarlo_results = QtWidgets.QLabel("Resultados aparecerán aquí tras ejecutar la simulación.")
        self.montecarlo_results.setWordWrap(True)
        
        self.montecarlo_figure = Figure()
        self.montecarlo_canvas = FigureCanvas(self.montecarlo_figure)
        
        layout.addLayout(risk_form)
        layout.addWidget(btn_run_montecarlo)
        layout.addWidget(self.montecarlo_results)
        layout.addWidget(self.montecarlo_canvas)
        tab.setLayout(layout)
        self.tabs.addTab(tab, "Análisis de Riesgos")

    def load_default_values(self):
        self.load_mining_costs()
        self.load_metal_prices()
        self.costo_otros.setText("10")
        self.planta_cost.setText("5")
        self.recuperacion_mineral.setText("90")
        self.recuperacion_metalurgica.setText("95")
        self.maquila.setText("5")
        self.price_table.setCurrentCell(0, 0)
        self.update_scenario_combo()

    def load_mining_costs(self):
        methods = [
            ["BC gravitational", 5, 1, 2, 0, 2, 10],
            ["BC LHD", 4, 1, 5, 0, 2, 12],
            ["BC MSGE", 6, 1, 3, 0, 2, 12],
            ["Sublevel Caving", 5, 3, 5, 0, 2, 15],
            ["Open Stopping", 5, 3, 5, 0, 3, 16],
            ["Room&Pillar", 5, 4, 5, 0, 3, 17],
            ["Shrinkage", 5, 4, 5, 0, 3, 17],
            ["LongWall/ShortWall", 6, 3, 5, 0, 3, 17],
            ["Filled Stopes", 6, 4, 6, 2, 4, 22],
            ["Filed Room&Pillar", 7, 5, 6, 2, 4, 24],
            ["Cut&Fill", 7, 5, 6, 2, 4, 24]
        ]
        self.cost_table.setRowCount(len(methods))
        for i, row in enumerate(methods):
            for j, value in enumerate(row):
                item = QtWidgets.QTableWidgetItem(str(value))
                self.cost_table.setItem(i, j, item)
        self.selected_method.addItems([m[0] for m in methods])

    def load_metal_prices(self):
        metals = [
            ["Oro", 2000], ["Plata", 22], ["Cobre", 4], ["Zinc", 1.35],
            ["Plomo", 1], ["Aluminio", 1.1], ["Niquel", 11], ["Estaño", 11],
            ["Hierro", 120]
        ]
        self.price_table.setRowCount(len(metals))
        for i, (metal, price) in enumerate(metals):
            self.price_table.setItem(i, 0, QtWidgets.QTableWidgetItem(metal))
            self.price_table.setItem(i, 1, QtWidgets.QTableWidgetItem(str(price)))

    def update_scenario_combo(self):
        self.scenario_combo.clear()
        self.scenario_combo.addItem("Seleccionar escenario...")
        if hasattr(self.parent(), 'tonelaje_scenarios'):
            scenarios = self.parent().tonelaje_scenarios
            for scenario in scenarios:
                self.scenario_combo.addItem(f"{scenario['cutoff']:.2f}%")

    def set_scenario(self, text):
        if text and text != "Seleccionar escenario...":
            cutoff_str = text.strip('%')
            cutoff_value = float(cutoff_str)
            for scenario in self.parent().tonelaje_scenarios:
                if scenario['cutoff'] == cutoff_value:
                    production_anual = (scenario['tonelaje'] * 1e6) / int(self.project_years.text() or 10)
                    self.production.setText(str(int(production_anual)))
                    self.average_grade = scenario['ley_media']
                    if not self.risk_params["Producción Anual"]["mode"].text():
                        self.risk_params["Producción Anual"]["mode"].setText(str(int(production_anual)))
                    break
        else:
            self.average_grade = None

    def calculate_values(self):
        try:
            if not self.scenario_combo.currentText() or self.scenario_combo.currentText() == "Seleccionar escenario...":
                raise ValueError("Debe seleccionar un escenario de tonelaje-ley")
            cm = self.get_current_mining_cost()
            cp = float(self.planta_cost.text() or 0)
            co_percentage = float(self.costo_otros.text() or 0)
            co = co_percentage / 100 * (cm + cp)
            recovery_mineral = float(self.recuperacion_mineral.text() or 0) / 100
            recovery_metalurgica = float(self.recuperacion_metalurgica.text() or 0) / 100
            recovery_total = recovery_mineral * recovery_metalurgica
            price_metal = self.get_metal_price()
            if price_metal is None:
                raise ValueError("Error en precio del metal")
            tcrc_percentage = float(self.maquila.text() or 0) / 100
            tcrc = tcrc_percentage * price_metal
            pmetal = price_metal - tcrc
            production_anual = float(self.production.text() or 0)
            average_grade = self.average_grade
            metal_produced = (production_anual * average_grade / 100) * recovery_total
            total_direct_cost = (cm + cp + co) * production_anual
            G_and_A = float(self.G_and_A_input.text() or 0)
            depreciation = float(self.depreciation_input.text() or 0)
            sustaining_capital = float(self.sustaining_capital_input.text() or 0)
            exploration_expenses = float(self.exploration_expenses_input.text() or 0)
            total_costs = total_direct_cost + G_and_A + depreciation + sustaining_capital + exploration_expenses
            revenue = pmetal * metal_produced
            flujo_anual = revenue - total_costs
            inversion_inicial = float(self.inversion_inicial.text() or 0)
            tasa = float(self.discount_rate.text() or 0) / 100
            años = int(self.project_years.text() or 10)
            van = -inversion_inicial
            for año in range(1, años + 1):
                van += flujo_anual / ((1 + tasa) ** año)
            if metal_produced > 0:
                AISC = total_costs / metal_produced
            else:
                AISC = "N/A"
            self.cutoff_grade_result.setText(f"{average_grade:.4f}%")
            self.van_result.setText(f"US$ {van:,.2f}")
            self.AISC_result.setText(f"US$ {AISC:,.2f}" if isinstance(AISC, (float, int)) else "N/A")
            self.results = {
                "Ley de Corte (%)": average_grade,
                "VAN (US$)": van,
                "Costo Minero (US$/ton)": cm,
                "Costo Planta (US$/ton)": cp,
                "Costo Otros (%)": co_percentage,
                "Recuperación Total (%)": recovery_total * 100,
                "Precio Metal (US$)": price_metal,
                "Producción Anual (ton)": production_anual,
                "Inversión Inicial (US$)": inversion_inicial,
                "G&A Costs (US$ anuales)": G_and_A,
                "Depreciación (US$ anuales)": depreciation,
                "Capital Sostenible (US$ anuales)": sustaining_capital,
                "Gastos Exploración (US$ anuales)": exploration_expenses,
                "AISC (US$/ton metal)": AISC if isinstance(AISC, (float, int)) else AISC
            }
        except ValueError as ve:
            self.show_error(f"Error de entrada: {str(ve)}")
        except Exception as e:
            self.show_error(f"Error en cálculo: {str(e)}")

    def run_montecarlo(self):
        try:
            if not self.scenario_combo.currentText() or self.scenario_combo.currentText() == "Seleccionar escenario...":
                raise ValueError("Debe seleccionar un escenario de tonelaje-ley en la pestaña de Análisis")
            num_simulations = int(self.num_simulations.text() or 1000)
            años = int(self.project_years.text() or 10)
            tasa = float(self.discount_rate.text() or 10) / 100
            inversion_inicial = float(self.inversion_inicial.text() or 0)
            G_and_A = float(self.G_and_A_input.text() or 0)
            depreciation = float(self.depreciation_input.text() or 0)
            sustaining_capital = float(self.sustaining_capital_input.text() or 0)
            exploration_expenses = float(self.exploration_expenses_input.text() or 0)
            co_percentage = float(self.costo_otros.text() or 0)
            maquila_percentage = float(self.maquila.text() or 0) / 100
            price_metal_dist = stats.triang(
                c=(float(self.risk_params["Precio Metal"]["mode"].text()) - float(self.risk_params["Precio Metal"]["min"].text())) / 
                  (float(self.risk_params["Precio Metal"]["max"].text()) - float(self.risk_params["Precio Metal"]["min"].text())),
                loc=float(self.risk_params["Precio Metal"]["min"].text()),
                scale=float(self.risk_params["Precio Metal"]["max"].text()) - float(self.risk_params["Precio Metal"]["min"].text())
            )
            cm_dist = stats.triang(
                c=(float(self.risk_params["Costo Minero"]["mode"].text()) - float(self.risk_params["Costo Minero"]["min"].text())) / 
                  (float(self.risk_params["Costo Minero"]["max"].text()) - float(self.risk_params["Costo Minero"]["min"].text())),
                loc=float(self.risk_params["Costo Minero"]["min"].text()),
                scale=float(self.risk_params["Costo Minero"]["max"].text()) - float(self.risk_params["Costo Minero"]["min"].text())
            )
            cp_dist = stats.triang(
                c=(float(self.risk_params["Costo Planta"]["mode"].text()) - float(self.risk_params["Costo Planta"]["min"].text())) / 
                  (float(self.risk_params["Costo Planta"]["max"].text()) - float(self.risk_params["Costo Planta"]["min"].text())),
                loc=float(self.risk_params["Costo Planta"]["min"].text()),
                scale=float(self.risk_params["Costo Planta"]["max"].text()) - float(self.risk_params["Costo Planta"]["min"].text())
            )
            production_dist = stats.triang(
                c=(float(self.risk_params["Producción Anual"]["mode"].text()) - float(self.risk_params["Producción Anual"]["min"].text())) / 
                  (float(self.risk_params["Producción Anual"]["max"].text()) - float(self.risk_params["Producción Anual"]["min"].text())),
                loc=float(self.risk_params["Producción Anual"]["min"].text()),
                scale=float(self.risk_params["Producción Anual"]["max"].text()) - float(self.risk_params["Producción Anual"]["min"].text())
            )
            recovery_dist = stats.triang(
                c=(float(self.risk_params["Recuperación Total"]["mode"].text()) - float(self.risk_params["Recuperación Total"]["min"].text())) / 
                  (float(self.risk_params["Recuperación Total"]["max"].text()) - float(self.risk_params["Recuperación Total"]["min"].text())),
                loc=float(self.risk_params["Recuperación Total"]["min"].text()),
                scale=float(self.risk_params["Recuperación Total"]["max"].text()) - float(self.risk_params["Recuperación Total"]["min"].text())
            )
            vans = []
            aisccs = []
            for _ in range(num_simulations):
                price_metal = price_metal_dist.rvs()
                cm = cm_dist.rvs()
                cp = cp_dist.rvs()
                production_anual = production_dist.rvs()
                recovery_total = recovery_dist.rvs() / 100
                co = co_percentage / 100 * (cm + cp)
                tcrc = maquila_percentage * price_metal
                pmetal = price_metal - tcrc
                metal_produced = (production_anual * self.average_grade / 100) * recovery_total
                total_direct_cost = (cm + cp + co) * production_anual
                total_costs = total_direct_cost + G_and_A + depreciation + sustaining_capital + exploration_expenses
                revenue = pmetal * metal_produced
                flujo_anual = revenue - total_costs
                van = -inversion_inicial
                for año in range(1, años + 1):
                    van += flujo_anual / ((1 + tasa) ** año)
                vans.append(van)
                if metal_produced > 0:
                    aisc = total_costs / metal_produced
                    aisccs.append(aisc)
            van_mean = np.mean(vans)
            van_std = np.std(vans)
            van_p10 = np.percentile(vans, 10)
            van_p50 = np.percentile(vans, 50)
            van_p90 = np.percentile(vans, 90)
            prob_positive_van = np.mean(np.array(vans) > 0) * 100
            aisc_mean = np.mean(aisccs)
            result_text = (
                f"<b>Resultados Monte Carlo ({num_simulations} simulaciones):</b><br>"
                f"VAN Promedio: US$ {van_mean:,.2f}<br>"
                f"Desviación Estándar VAN: US$ {van_std:,.2f}<br>"
                f"VAN P10: US$ {van_p10:,.2f}<br>"
                f"VAN P50: US$ {van_p50:,.2f}<br>"
                f"VAN P90: US$ {van_p90:,.2f}<br>"
                f"Probabilidad VAN > 0: {prob_positive_van:.2f}%<br>"
                f"AISC Promedio: US$ {aisc_mean:,.2f}/ton metal"
            )
            self.montecarlo_results.setText(result_text)
            self.montecarlo_figure.clear()
            ax = self.montecarlo_figure.add_subplot(111)
            ax.hist(vans, bins=50, color='blue', alpha=0.7, density=True)
            ax.set_xlabel("VAN (US$)")
            ax.set_ylabel("Densidad de Probabilidad")
            ax.set_title("Distribución del VAN - Monte Carlo")
            ax.axvline(0, color='red', linestyle='--', label='VAN = 0')
            ax.legend()
            self.montecarlo_canvas.draw()
            self.montecarlo_data = pd.DataFrame({"VAN": vans, "AISC": aisccs})
        except ValueError as ve:
            self.show_error(f"Error de entrada: {str(ve)}")
        except Exception as e:
            self.show_error(f"Error en Monte Carlo: {str(e)}")

    def get_current_mining_cost(self):
        index = self.selected_method.currentIndex()
        return float(self.cost_table.item(index, 6).text())

    def get_metal_price(self):
        try:
            selected_row = self.price_table.currentRow()
            if selected_row == -1:
                raise ValueError("Debe seleccionar un metal en la tabla de precios")
            metal_item = self.price_table.item(selected_row, 0)
            price_item = self.price_table.item(selected_row, 1)
            if not metal_item or not price_item:
                raise ValueError("Datos incompletos en la tabla de precios")
            selected_metal = metal_item.text()
            price = float(price_item.text())
            conversion_factors = {
                "Oro": 32.15074, "Plata": 32.15074, "Cobre": 2204.62,
                "Zinc": 2204.62, "Plomo": 2204.62, "Aluminio": 2204.62,
                "Niquel": 1, "Estaño": 1, "Hierro": 1
            }
            return price * conversion_factors.get(selected_metal, 1)
        except Exception as e:
            self.show_error(f"Error en precio: {str(e)}")
            return None

    def export_to_excel(self):
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Exportar a Excel", "", "Excel (*.xlsx)")
        if file_path:
            try:
                with pd.ExcelWriter(file_path) as writer:
                    if hasattr(self, 'results'):
                        df_results = pd.DataFrame(list(self.results.items()), columns=["Parámetro", "Valor"])
                        df_results.to_excel(writer, sheet_name="Análisis Determinístico", index=False)
                    if hasattr(self, 'montecarlo_data'):
                        self.montecarlo_data.to_excel(writer, sheet_name="Monte Carlo", index=False)
                QtWidgets.QMessageBox.information(self, "Éxito", "Resultados exportados a Excel!")
            except Exception as e:
                self.show_error(f"No se pudo exportar:\n{str(e)}")

    def show_error(self, message):
        QtWidgets.QMessageBox.critical(self, "Error", message)

class AdvancedBlockViewer(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("MineChart - Visualizador de Bloques")
        self.setGeometry(100, 100, 1600, 1000)
        self.dfs = {}
        self.current_file = None
        self.df = None
        self.grid = None
        self.active_filters = {}
        self.base_grid = None
        self.transparency = 80
        self.column_mappings = {}
        self.current_variable = 'ley'
        self.custom_colors = {'min': None, 'mid': None, 'max': None}
        self.file_path = None
        self.selected_cell_index = None
        self.cmap = 'viridis'
        self.predefined_filters = []
        self.tonelaje_scenarios = []
        self.plotter = QtInteractor(self)
        self.axes_visible = True
        self.grid_visible = False
        self.grid_actor = None
        self.apply_global_style()
        self.setup_ui()
        self.setup_view_controls()
        self.connect_events()

    def apply_global_style(self):
        style = """
            QWidget { font-family: 'Segoe UI', Arial, sans-serif; font-size: 12px; background-color: #f0f0f0; }
            QToolBar { background-color: #f0f0f0; border: none; spacing: 2px; }
            QToolButton { 
                background-color: transparent; 
                border: none; 
                padding: 4px; 
                margin: 2px; 
            }
            QToolButton:hover { 
                background-color: #e0e0e0; 
                border-radius: 4px; 
            }
            QToolButton:pressed { 
                background-color: #d0d0d0; 
            }
            QGroupBox { border: 1px solid #cccccc; border-radius: 4px; margin-top: 20px; padding-top: 10px; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 3px; }
            QLineEdit, QComboBox, QSlider { background-color: #ffffff; border: 1px solid #cccccc; 
                                          padding: 4px; border-radius: 3px; }
            QTabWidget::pane { border: 1px solid #cccccc; background: #ffffff; }
            QTabBar::tab { background: #f0f0f0; border: 1px solid #cccccc; padding: 8px 12px; margin-right: 2px; }
            QTabBar::tab:selected { background: #ffffff; border-bottom-color: #ffffff; }
        """
        self.setStyleSheet(style)

    def setup_ui(self):
        central_widget = QtWidgets.QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QtWidgets.QVBoxLayout(central_widget)
        toolbar = QtWidgets.QToolBar("Herramientas", self)
        toolbar.setIconSize(QtCore.QSize(24, 24))
        self.addToolBar(QtCore.Qt.TopToolBarArea, toolbar)
        action_load = QtWidgets.QAction(QtGui.QIcon("icons/open.png"), "Cargar archivo", self)
        action_load.triggered.connect(self.load_file)
        action_edit = QtWidgets.QAction(QtGui.QIcon("icons/edit.png"), "Editar Columnas", self)
        action_edit.triggered.connect(self.edit_columns)
        action_stats = QtWidgets.QAction(QtGui.QIcon("icons/stats.png"), "Reporte Estadístico", self)
        action_stats.triggered.connect(self.show_statistical_report)
        action_tonnage = QtWidgets.QAction(QtGui.QIcon("icons/tonnage.png"), "Curva Ton-Ley", self)
        action_tonnage.triggered.connect(self.show_tonnage_curve)
        action_economic = QtWidgets.QAction(QtGui.QIcon("icons/economic.png"), "Análisis Económico", self)
        action_economic.triggered.connect(self.show_economic_analysis)
        action_view_data = QtWidgets.QAction(QtGui.QIcon("icons/table.png"), "Ver Datos", self)
        action_view_data.triggered.connect(self.show_data_table)
        action_save_template = QtWidgets.QAction(QtGui.QIcon("icons/save-template.png"), "Guardar Plantilla", self)
        action_save_template.triggered.connect(self.save_template)
        action_load_template = QtWidgets.QAction(QtGui.QIcon("icons/load-template.png"), "Cargar Plantilla", self)
        action_load_template.triggered.connect(self.load_template)
        toolbar.addAction(action_load)
        toolbar.addAction(action_edit)
        toolbar.addAction(action_view_data)
        toolbar.addSeparator()
        toolbar.addAction(action_stats)
        toolbar.addAction(action_tonnage)
        toolbar.addAction(action_economic)
        toolbar.addSeparator()
        toolbar.addAction(action_save_template)
        toolbar.addAction(action_load_template)
        toolbar.addSeparator()
        toolbar.addWidget(QtWidgets.QLabel("Archivo:"))
        self.file_combo = QtWidgets.QComboBox()
        self.file_combo.currentTextChanged.connect(self.switch_file)
        toolbar.addWidget(self.file_combo)
        bottom_panel = QtWidgets.QHBoxLayout()
        self.plotter.set_background("#2E2E2E")
        self.right_panel = QtWidgets.QWidget()
        right_layout = QtWidgets.QVBoxLayout(self.right_panel)
        filter_group = QtWidgets.QGroupBox("Filtrar")
        filter_layout = QtWidgets.QVBoxLayout()
        filter_controls = QtWidgets.QHBoxLayout()
        self.column_combo = QtWidgets.QComboBox()
        self.filter_input = QtWidgets.QLineEdit()
        self.filter_input.setPlaceholderText("Operador y valor (ej: >0.5, 0.5-0.7)")
        self.color_button = QtWidgets.QPushButton("Seleccionar color")
        self.btn_add_filter = QtWidgets.QPushButton("Añadir filtro")
        filter_controls.addWidget(self.column_combo)
        filter_controls.addWidget(self.filter_input)
        filter_layout.addLayout(filter_controls)
        filter_layout.addWidget(self.color_button)
        filter_layout.addWidget(self.btn_add_filter)
        filter_layout.addWidget(QtWidgets.QLabel("Filtros predefinidos:"))
        self.predefined_combo = QtWidgets.QComboBox()
        self.predefined_combo.addItem("Seleccionar filtro...")
        self.btn_save_filter = QtWidgets.QPushButton("Guardar Filtro")
        self.btn_save_filter.clicked.connect(self.save_predefined_filter)
        self.predefined_combo.currentTextChanged.connect(self.apply_predefined_filter)
        filter_layout.addWidget(self.predefined_combo)
        filter_layout.addWidget(self.btn_save_filter)
        self.filters_list = QtWidgets.QListWidget()
        filter_layout.addWidget(QtWidgets.QLabel("Filtros activos:"))
        filter_layout.addWidget(self.filters_list)
        filter_group.setLayout(filter_layout)
        view_group = QtWidgets.QGroupBox("Controles de Vista")
        view_layout = QtWidgets.QGridLayout()
        self.btn_reset = QtWidgets.QPushButton("Reiniciar vista")
        self.btn_axes = QtWidgets.QPushButton("Mostrar ejes")
        self.btn_grid = QtWidgets.QPushButton("Mostrar grilla")
        for btn, color in [(self.btn_reset, "#4062a0"), (self.btn_axes, "#4062a0"), (self.btn_grid, "#4062a0")]:
            btn.setStyleSheet(self.get_button_style(color))
        view_layout.addWidget(self.btn_reset, 0, 0)
        view_layout.addWidget(self.btn_axes, 0, 1)
        view_layout.addWidget(self.btn_grid, 1, 0)
        view_layout.addWidget(QtWidgets.QLabel("ColorMap:"), 1, 1)
        self.cmap_combo = QtWidgets.QComboBox()
        self.cmap_combo.addItems(['viridis', 'plasma', 'inferno', 'magma', 'cividis', 'coolwarm', 'rainbow'])
        view_layout.addWidget(self.cmap_combo, 2, 0, 1, 2)
        view_group.setLayout(view_layout)
        transparency_group = QtWidgets.QGroupBox("Transparencia Base")
        transparency_layout = QtWidgets.QVBoxLayout()
        self.transparency_slider = QtWidgets.QSlider(QtCore.Qt.Horizontal)
        self.transparency_slider.setRange(0, 100)
        self.transparency_slider.setValue(self.transparency)
        self.transparency_label = QtWidgets.QLabel(f"Transparencia: {self.transparency}%")
        transparency_layout.addWidget(self.transparency_label)
        transparency_layout.addWidget(self.transparency_slider)
        transparency_group.setLayout(transparency_layout)
        gradient_group = QtWidgets.QGroupBox("Gradiente Personalizado")
        gradient_layout = QtWidgets.QHBoxLayout()
        self.btn_min_color = QtWidgets.QPushButton("Mínimo")
        self.btn_mid_color = QtWidgets.QPushButton("Medio")
        self.btn_max_color = QtWidgets.QPushButton("Máximo")
        for btn in [self.btn_min_color, self.btn_mid_color, self.btn_max_color]:
            btn.setStyleSheet("background-color: gray; color: white;")
        gradient_layout.addWidget(self.btn_min_color)
        gradient_layout.addWidget(self.btn_mid_color)
        gradient_layout.addWidget(self.btn_max_color)
        gradient_group.setLayout(gradient_layout)
        right_layout.addWidget(filter_group)
        right_layout.addWidget(view_group)
        right_layout.addWidget(transparency_group)
        right_layout.addWidget(gradient_group)
        right_layout.addStretch()
        bottom_panel.addWidget(self.plotter, 4)
        bottom_panel.addWidget(self.right_panel, 1)
        main_layout.addLayout(bottom_panel)

    def setup_view_controls(self):
        self.plotter.enable_terrain_style()
        self.plotter.enable_anti_aliasing()
        self.plotter.show_axes()
        self.axes_visible = True
        self.grid_actor = None
        self.grid_visible = False

    def connect_events(self):
        self.btn_add_filter.clicked.connect(self.add_filter)
        self.color_button.clicked.connect(self.select_color)
        self.btn_reset.clicked.connect(self.reset_view)
        self.btn_axes.clicked.connect(self.toggle_axes)
        self.btn_grid.clicked.connect(self.toggle_grid)
        self.transparency_slider.valueChanged.connect(self.update_transparency)
        self.column_combo.currentTextChanged.connect(self.update_current_variable)
        self.plotter.iren.add_observer(vtkCommand.MouseMoveEvent, self.on_mouse_move)
        self.plotter.iren.add_observer(vtkCommand.LeftButtonPressEvent, self.on_click)
        self.btn_min_color.clicked.connect(lambda: self.select_gradient_color('min'))
        self.btn_mid_color.clicked.connect(lambda: self.select_gradient_color('mid'))
        self.btn_max_color.clicked.connect(lambda: self.select_gradient_color('max'))
        self.cmap_combo.currentTextChanged.connect(self.update_cmap)

    def reset_view(self):
        self.plotter.reset_camera()
        self.plotter.view_isometric()
        self.plotter.render()

    def toggle_axes(self):
        if self.axes_visible:
            self.plotter.hide_axes()
            self.axes_visible = False
        else:
            self.plotter.show_axes()
            self.axes_visible = True
        self.btn_axes.setStyleSheet(
            self.get_button_style("#4062a0") if self.axes_visible else self.get_button_style("#808080")
        )
        self.plotter.render()

    def toggle_grid(self):
        if self.grid_visible:
            if self.grid_actor is not None:
                self.plotter.remove_actor(self.grid_actor)
                self.grid_actor = None
            self.grid_visible = False
        else:
            self.grid_actor = self.plotter.show_grid()
            self.grid_visible = True
        self.btn_grid.setStyleSheet(
            self.get_button_style("#4062a0") if self.grid_visible else self.get_button_style("#808080")
        )
        self.plotter.render()

    def on_click(self, obj, event):
        interactor = self.plotter.iren.get_interactor_style()
        x, y = self.plotter.iren.get_event_position()
        picker = vtkCellPicker()
        picker.Pick(x, y, 0, self.plotter.renderer)
        cell_id = picker.GetCellId()
        if cell_id != -1:
            self.selected_cell_index = cell_id
            self.highlight_selected_cell()
            self.show_cell_info(cell_id)
        else:
            self.clear_selection()

    def on_mouse_move(self, obj, event):
        pass

    def highlight_selected_cell(self):
        if self.selected_cell_index is not None and self.grid is not None:
            highlight_grid = self.grid.extract_cells([self.selected_cell_index])
            self.plotter.add_mesh(
                highlight_grid, color='yellow', opacity=1.0, name='highlight'
            )
            self.plotter.render()

    def clear_selection(self):
        self.selected_cell_index = None
        self.plotter.remove_actor('highlight')
        self.plotter.render()

    def show_cell_info(self, cell_id):
        if self.df is not None and cell_id < len(self.df):
            info = self.df.iloc[cell_id]
            message = "\n".join([f"{col}: {val}" for col, val in info.items()])
            QtWidgets.QMessageBox.information(self, "Información de Bloque", message)

    def load_file(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Abrir archivo", "", "Archivos soportados (*.xlsx *.csv)"
        )
        if file_path:
            try:
                df = pd.read_excel(file_path) if file_path.endswith('.xlsx') else pd.read_csv(file_path)
                required = ['x0', 'y0', 'z0', 'xinc', 'yinc', 'zinc', 'ley']
                if not all(col in df.columns for col in required):
                    dlg = ColumnMapperDialog(df.columns.tolist(), self)
                    if dlg.exec_() == QtWidgets.QDialog.Accepted:
                        self.column_mappings = dlg.mappings
                        df.rename(columns={v: k for k, v in dlg.mappings.items()}, inplace=True)
                    else:
                        return
                self.dfs[file_path] = df
                self.file_path = file_path
                self.current_file = file_path
                self.df = self.dfs[self.current_file]
                self.create_3d_model()
                self.update_column_combo()
                self.update_visualization()
                self.update_file_combo()
            except Exception as e:
                self.show_error(f"Error al cargar archivo:\n{str(e)}")

    def update_file_combo(self):
        self.file_combo.clear()
        self.file_combo.addItems(list(self.dfs.keys()))
        if self.current_file:
            self.file_combo.setCurrentText(self.current_file)

    def switch_file(self, file_path):
        if file_path in self.dfs:
            self.current_file = file_path
            self.df = self.dfs[self.current_file]
            self.file_path = self.current_file
            self.create_3d_model()
            self.update_column_combo()
            self.filters_list.clear()
            self.update_visualization()

    def update_column_combo(self):
        self.column_combo.clear()
        if self.df is not None:
            numeric_cols = self.df.select_dtypes(include=np.number).columns.tolist()
            self.column_combo.addItems(numeric_cols)
            if self.current_variable in numeric_cols:
                self.column_combo.setCurrentText(self.current_variable)
            elif 'ley' in numeric_cols:
                self.current_variable = 'ley'
                self.column_combo.setCurrentText('ley')
            else:
                self.current_variable = numeric_cols[0] if numeric_cols else None
                self.column_combo.setCurrentText(self.current_variable if self.current_variable else '')

    def edit_columns(self):
        if self.df is not None:
            dialog = FormulaEditor(self.df, self.file_path, self)
            if dialog.exec_() == QtWidgets.QDialog.Accepted:
                self.df = dialog.df
                self.create_3d_model()
                self.update_column_combo()
                self.filters_list.clear()
                self.update_visualization()
        else:
            self.show_error("Cargue un archivo primero")

    def create_3d_model(self):
        if self.df is None:
            return
        x0 = self.df['x0'].astype(float)
        y0 = self.df['y0'].astype(float)
        z0 = self.df['z0'].astype(float)
        xinc = self.df['xinc'].astype(float)
        yinc = self.df['yinc'].astype(float)
        zinc = self.df['zinc'].astype(float)
        points = []
        cells = []
        cell_types = []
        for i in range(len(self.df)):
            x, y, z = x0[i], y0[i], z0[i]
            dx, dy, dz = xinc[i], yinc[i], zinc[i]
            vertices = [
                (x, y, z),
                (x + dx, y, z),
                (x + dx, y + dy, z),
                (x, y + dy, z),
                (x, y, z + dz),
                (x + dx, y, z + dz),
                (x + dx, y + dy, z + dz),
                (x, y + dy, z + dz)
            ]
            base_idx = len(points)
            points.extend(vertices)
            cells.append(8)
            cells.extend([base_idx + j for j in range(8)])
            cell_types.append(pv.CellType.HEXAHEDRON)
        points = np.array(points)
        cells = np.array(cells)
        self.grid = pv.UnstructuredGrid(cells, cell_types, points)
        for col in self.df.select_dtypes(include=np.number).columns:
            self.grid.cell_data[col] = self.df[col].astype(float).values
        self.base_grid = self.grid.copy()

    def update_current_variable(self, var):
        if var:
            self.current_variable = var
            self.update_visualization()

    def update_transparency(self, value):
        self.transparency = value
        self.transparency_label.setText(f"Transparencia: {self.transparency}%")
        self.update_visualization()

    def select_color(self):
        color = QtWidgets.QColorDialog.getColor()
        if color.isValid():
            self.color_button.setStyleSheet(f"background-color: {color.name()}; border: none;")
            self.add_filter()

    def add_filter(self):
        column = self.column_combo.currentText()
        filter_text = self.filter_input.text().strip()
        if not column or not filter_text:
            self.show_error("Seleccione una columna y especifique un filtro")
            return
        try:
            if '-' in filter_text:
                operator = 'range'
                low, high = map(float, filter_text.split('-'))
                threshold = (low, high)
            else:
                operator = None
                for op in ['>', '<', '>=', '<=', '==']:
                    if filter_text.startswith(op):
                        operator = op
                        threshold = float(filter_text[len(op):])
                        break
                if not operator:
                    raise ValueError("Operador no válido. Use >, <, >=, <=, == o un rango (ej: 0.5-0.7)")
            color = self.color_button.palette().button().color().name()
            filter_data = {
                'column': column,
                'operator': operator,
                'threshold': threshold,
                'color': color,
                'transparency': 100,
                'active': True
            }
            item = QtWidgets.QListWidgetItem()
            widget = QtWidgets.QWidget()
            layout = QtWidgets.QHBoxLayout()
            chk = QtWidgets.QCheckBox("Activo")
            chk.setChecked(True)
            label_text = f"{column} {operator} {threshold}" if operator != 'range' else f"{column} {threshold[0]} - {threshold[1]}"
            label = QtWidgets.QLabel(label_text)
            color_label = QtWidgets.QLabel()
            color_label.setStyleSheet(f"background-color: {color}; border-radius: 4px;")
            color_label.setFixedSize(24, 24)
            transparency_slider = QtWidgets.QSlider(QtCore.Qt.Horizontal)
            transparency_slider.setRange(0, 100)
            transparency_slider.setValue(100)
            transparency_slider.setFixedWidth(80)
            btn_remove = QtWidgets.QPushButton("X")
            btn_remove.setStyleSheet(self.get_button_style("#FF5252"))
            btn_remove.setFixedSize(24, 24)
            layout.addWidget(chk)
            layout.addWidget(label)
            layout.addWidget(color_label)
            layout.addWidget(transparency_slider)
            layout.addWidget(btn_remove)
            layout.setContentsMargins(0, 0, 0, 0)
            widget.setLayout(layout)
            item.setSizeHint(widget.sizeHint())
            item.filter_data = filter_data
            transparency_slider.valueChanged.connect(lambda v: self.update_filter_transparency(item-IMPROVE_CODE_HERE, v))
            btn_remove.clicked.connect(lambda: self.remove_filter(item))
            chk.stateChanged.connect(lambda s: self.toggle_filter(item, s))
            self.filters_list.addItem(item)
            self.filters_list.setItemWidget(item, widget)
            self.apply_filter(filter_data)
            self.update_visualization()
        except ValueError as e:
            self.show_error(str(e))
        except Exception as e:
            self.show_error(f"Error al agregar filtro: {str(e)}")

    def save_predefined_filter(self):
        column = self.column_combo.currentText()
        filter_text = self.filter_input.text()
        color = self.color_button.palette().button().color().name()
        if not column or not filter_text:
            QtWidgets.QMessageBox.warning(self, "Error", "Complete los campos para guardar el filtro.")
            return
        filter_name = QtWidgets.QInputDialog.getText(self, "Nombre del Filtro", "Ingrese un nombre para el filtro:")[0]
        if filter_name:
            try:
                if '-' in filter_text:
                    operator = 'range'
                    low, high = map(float, filter_text.split('-'))
                    threshold = (low, high)
                else:
                    operator = None
                    for op in ['>', '<', '>=', '<=', '==']:
                        if filter_text.startswith(op):
                            operator = op
                            threshold = float(filter_text[len(op):])
                            break
                    if not operator:
                        raise ValueError("Operador no válido.")
                self.predefined_filters.append({
                    "name": filter_name,
                    "column": column,
                    "filter_text": filter_text,
                    "color": color,
                    "operator": operator,
                    "threshold": threshold
                })
                self.predefined_combo.addItem(filter_name)
                QtWidgets.QMessageBox.information(self, "Éxito", "Filtro guardado!")
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "Error", f"Error al guardar filtro:\n{str(e)}")

    def apply_predefined_filter(self, filter_name):
        if filter_name == "Seleccionar filtro...":
            return
        for filter_data in self.predefined_filters:
            if filter_data["name"] == filter_name:
                self.column_combo.setCurrentText(filter_data["column"])
                self.filter_input.setText(filter_data["filter_text"])
                self.color_button.setStyleSheet(f"background-color: {filter_data['color']}; border: none;")
                filter_data_full = {
                    'column': filter_data["column"],
                    'operator': filter_data["operator"],
                    'threshold': filter_data["threshold"],
                    'color': filter_data["color"],
                    'transparency': 100,
                    'active': True
                }
                item = QtWidgets.QListWidgetItem()
                widget = QtWidgets.QWidget()
                layout = QtWidgets.QHBoxLayout()
                chk = QtWidgets.QCheckBox("Activo")
                chk.setChecked(True)
                label_text = f"{filter_data['column']} {filter_data['operator']} {filter_data['threshold']}" if filter_data['operator'] != 'range' else f"{filter_data['column']} {filter_data['threshold'][0]} - {filter_data['threshold'][1]}"
                label = QtWidgets.QLabel(label_text)
                color_label = QtWidgets.QLabel()
                color_label.setStyleSheet(f"background-color: {filter_data['color']}; border-radius: 4px;")
                color_label.setFixedSize(24, 24)
                transparency_slider = QtWidgets.QSlider(QtCore.Qt.Horizontal)
                transparency_slider.setRange(0, 100)
                transparency_slider.setValue(100)
                transparency_slider.setFixedWidth(80)
                btn_remove = QtWidgets.QPushButton("X")
                btn_remove.setStyleSheet(self.get_button_style("#FF5252"))
                btn_remove.setFixedSize(24, 24)
                layout.addWidget(chk)
                layout.addWidget(label)
                layout.addWidget(color_label)
                layout.addWidget(transparency_slider)
                layout.addWidget(btn_remove)
                layout.setContentsMargins(0, 0, 0, 0)
                widget.setLayout(layout)
                item.setSizeHint(widget.sizeHint())
                item.filter_data = filter_data_full
                transparency_slider.valueChanged.connect(lambda v: self.update_filter_transparency(item, v))
                btn_remove.clicked.connect(lambda: self.remove_filter(item))
                chk.stateChanged.connect(lambda s: self.toggle_filter(item, s))
                self.filters_list.addItem(item)
                self.filters_list.setItemWidget(item, widget)
                self.apply_filter(filter_data_full)
                self.update_visualization()
                break

    def toggle_filter(self, item, state):
        item.filter_data['active'] = state == QtCore.Qt.Checked
        self.update_visualization()

    def remove_filter(self, item):
        row = self.filters_list.row(item)
        self.filters_list.takeItem(row)
        self.update_visualization()

    def update_filter_transparency(self, item, value):
        item.filter_data['transparency'] = value
        self.update_visualization()

    def apply_filter(self, filter_data):
        self.active_filters[filter_data['column']] = filter_data

    def select_gradient_color(self, position):
        color = QtWidgets.QColorDialog.getColor()
        if color.isValid():
            self.custom_colors[position] = color.name()
            getattr(self, f'btn_{position}_color').setStyleSheet(
                f"background-color: {color.name()}; color: white;"
            )
            self.update_cmap()

    def update_cmap(self, cmap_name=None):
        if cmap_name:
            self.cmap = cmap_name
        if all(self.custom_colors.values()):
            colors = [self.custom_colors['min'], self.custom_colors['mid'], self.custom_colors['max']]
            self.cmap = LinearSegmentedColormap.from_list('custom', colors)
            plt.register_cmap(cmap=self.cmap)
        self.update_visualization()

    def update_visualization(self):
        if self.grid is None or self.df is None:
            return
        self.plotter.clear()
        filtered_grids = {}
        remaining_cells = set(range(self.grid.n_cells))
        for filter_data in [item.filter_data for item in [self.filters_list.item(i) for i in range(self.filters_list.count())] if item.filter_data['active']]:
            column = filter_data['column']
            operator = filter_data['operator']
            threshold = filter_data['threshold']
            color = filter_data['color']
            transparency = filter_data['transparency'] / 100.0
            values = self.df[column].astype(float).values
            if operator == 'range':
                mask = (values >= threshold[0]) & (values <= threshold[1])
            else:
                mask = eval(f"values {operator} threshold")
            cell_indices = np.where(mask)[0]
            filtered_grid = self.grid.extract_cells(cell_indices)
            filtered_grids[column] = (filtered_grid, color, transparency)
            remaining_cells -= set(cell_indices)
        base_opacity = self.transparency / 100.0
        if remaining_cells:
            base_grid = self.grid.extract_cells(list(remaining_cells))
            self.plotter.add_mesh(
                base_grid,
                scalars=self.current_variable,
                cmap=self.cmap,
                opacity=base_opacity,
                show_scalar_bar=True,
                scalar_bar_args={'title': self.current_variable, 'vertical': True},
                name='base'
            )
        for column, (grid, color, opacity) in filtered_grids.items():
            self.plotter.add_mesh(
                grid,
                color=color,
                opacity=opacity,
                name=f'filter_{column}'
            )
        if self.selected_cell_index is not None:
            self.highlight_selected_cell()
        self.plotter.render()

    def show_statistical_report(self):
        if self.df is None:
            self.show_error("Cargue un archivo primero")
            return
        dialog = StatisticalReportDialog(self.df, self.current_variable, self)
        dialog.exec_()

    def show_tonnage_curve(self):
        if self.df is None:
            self.show_error("Cargue un archivo primero")
            return
        dialog = TonnageLeyDialog(self.df, self)
        dialog.exec_()

    def show_economic_analysis(self):
        if self.df is None:
            self.show_error("Cargue un archivo primero")
            return
        dialog = EconomicAnalysisDialog(self.df, self)
        dialog.exec_()

    def show_data_table(self):
        if self.df is None:
            self.show_error("Cargue un archivo primero")
            return
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Datos del Archivo")
        layout = QtWidgets.QVBoxLayout(dialog)
        table = QtWidgets.QTableWidget()
        table.setRowCount(len(self.df))
        table.setColumnCount(len(self.df.columns))
        table.setHorizontalHeaderLabels(self.df.columns)
        for i in range(len(self.df)):
            for j in range(len(self.df.columns)):
                item = QtWidgets.QTableWidgetItem(str(self.df.iloc[i, j]))
                table.setItem(i, j, item)
        layout.addWidget(table)
        dialog.resize(800, 600)
        dialog.exec_()

    def save_template(self):
        if not self.active_filters and not self.predefined_filters:
            QtWidgets.QMessageBox.warning(self, "Advertencia", "No hay filtros para guardar.")
            return
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Guardar Plantilla", "", "JSON (*.json)")
        if file_path:
            template = {
                'active_filters': [item.filter_data for item in [self.filters_list.item(i) for i in range(self.filters_list.count())]],
                'predefined_filters': self.predefined_filters,
                'column_mappings': self.column_mappings,
                'transparency': self.transparency,
                'cmap': self.cmap,
                'custom_colors': self.custom_colors
            }
            with open(file_path, 'w') as f:
                json.dump(template, f, indent=4)
            QtWidgets.QMessageBox.information(self, "Éxito", "Plantilla guardada correctamente!")

    def load_template(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Cargar Plantilla", "", "JSON (*.json)")
        if file_path:
            try:
                with open(file_path, 'r') as f:
                    template = json.load(f)
                self.filters_list.clear()
                self.predefined_filters = template.get('predefined_filters', [])
                self.column_mappings = template.get('column_mappings', {})
                self.transparency = template.get('transparency', 80)
                self.transparency_slider.setValue(self.transparency)
                self.transparency_label.setText(f"Transparencia: {self.transparency}%")
                self.cmap = template.get('cmap', 'viridis')
                self.cmap_combo.setCurrentText(self.cmap)
                self.custom_colors = template.get('custom_colors', {'min': None, 'mid': None, 'max': None})
                for pos, color in self.custom_colors.items():
                    if color:
                        getattr(self, f'btn_{pos}_color').setStyleSheet(
                            f"background-color: {color}; color: white;"
                        )
                for filter_data in template.get('active_filters', []):
                    item = QtWidgets.QListWidgetItem()
                    widget = QtWidgets.QWidget()
                    layout = QtWidgets.QHBoxLayout()
                    chk = QtWidgets.QCheckBox("Activo")
                    chk.setChecked(filter_data.get('active', True))
                    label_text = f"{filter_data['column']} {filter_data['operator']} {filter_data['threshold']}" if filter_data['operator'] != 'range' else f"{filter_data['column']} {filter_data['threshold'][0]} - {filter_data['threshold'][1]}"
                    label = QtWidgets.QLabel(label_text)
                    color_label = QtWidgets.QLabel()
                    color_label.setStyleSheet(f"background-color: {filter_data['color']}; border-radius: 4px;")
                    color_label.setFixedSize(24, 24)
                    transparency_slider = QtWidgets.QSlider(QtCore.Qt.Horizontal)
                    transparency_slider.setRange(0, 100)
                    transparency_slider.setValue(filter_data.get('transparency', 100))
                    transparency_slider.setFixedWidth(80)
                    btn_remove = QtWidgets.QPushButton("X")
                    btn_remove.setStyleSheet(self.get_button_style("#FF5252"))
                    btn_remove.setFixedSize(24, 24)
                    layout.addWidget(chk)
                    layout.addWidget(label)
                    layout.addWidget(color_label)
                    layout.addWidget(transparency_slider)
                    layout.addWidget(btn_remove)
                    layout.setContentsMargins(0, 0, 0, 0)
                    widget.setLayout(layout)
                    item.setSizeHint(widget.sizeHint())
                    item.filter_data = filter_data
                    transparency_slider.valueChanged.connect(lambda v, i=item: self.update_filter_transparency(i, v))
                    btn_remove.clicked.connect(lambda: self.remove_filter(item))
                    chk.stateChanged.connect(lambda s, i=item: self.toggle_filter(i, s))
                    self.filters_list.addItem(item)
                    self.filters_list.setItemWidget(item, widget)
                self.predefined_combo.clear()
                self.predefined_combo.addItem("Seleccionar filtro...")
                self.predefined_combo.addItems([f["name"] for f in self.predefined_filters])
                self.update_visualization()
                QtWidgets.QMessageBox.information(self, "Éxito", "Plantilla cargada correctamente!")
            except Exception as e:
                self.show_error(f"Error al cargar plantilla:\n{str(e)}")

    def get_button_style(self, color):
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border-radius: 4px;
                padding: 6px;
                min-width: 60px;
            }}
            QPushButton:hover {{
                background-color: {self.lighten_color(color, 20)};
            }}
            QPushButton:pressed {{
                background-color: {self.lighten_color(color, -20)};
            }}
        """

    def lighten_color(self, color, amount):
        c = QtGui.QColor(color)
        h, s, l, a = c.getHsl()
        l = max(0, min(255, l + amount))
        c.setHsl(h, s, l, a)
        return c.name()

    def show_error(self, message):
        QtWidgets.QMessageBox.critical(self, "Error", message)

if __name__ == '__main__':
    warnings.filterwarnings("ignore", category=DeprecationWarning)
    app = QtWidgets.QApplication(sys.argv)
    window = AdvancedBlockViewer()
    window.show()
    sys.exit(app.exec_())
